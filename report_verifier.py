import os
import sys

import openpyxl


class ReportVerifier(object):

    def __init__(self):
        self.dictionary_results = {}
        self.list_report_items = ["component", "iface", "total_report", "RESULT"]
        self.sheet_polyspace = "Summary"
        self.sheet_final_report = "Details"
        self.report_folder = os.path.join(os.getcwd(), "report")

    def get_results_polyspace_report(self, polyspace_report):
        try:
            polyspace_workbook = openpyxl.load_workbook(polyspace_report)
            active_sheet = polyspace_workbook[self.sheet_polyspace]
            # get numbers of row
            max_row = active_sheet.max_row
            '''
            A = compo
            B = kind compo
            I = get arg
            '''
            needed_columns = ["A", "B", "I"]
            start_row_needed = 3
            for i in range(start_row_needed, max_row + 1):
                # need to check if we already have the key
                full_component_name = active_sheet["{}{}".format(needed_columns[0], i)].value
                comp_name = self.get_component_name(full_component_name)
                comp_dict = {}
                # check if we already have the key
                if comp_name in self.dictionary_results:
                    if str(active_sheet["{}{}".format(needed_columns[1], i)].value).lower() == "component":
                        comp_dict[self.list_report_items[0]] = active_sheet["{}{}".format(needed_columns[2], i)].value
                    elif str(active_sheet["{}{}".format(needed_columns[1], i)].value).lower() == "iface":
                        comp_dict[self.list_report_items[1]] = active_sheet["{}{}".format(needed_columns[2], i)].value
                    self.dictionary_results[comp_name].update(comp_dict)
                else:
                    if str(active_sheet["{}{}".format(needed_columns[1], i)].value).lower() == "component":
                        comp_dict[self.list_report_items[0]] = active_sheet["{}{}".format(needed_columns[2], i)].value
                    elif str(active_sheet["{}{}".format(needed_columns[1], i)].value).lower() == "iface":
                        comp_dict[self.list_report_items[1]] = active_sheet["{}{}".format(needed_columns[2], i)].value
                    self.dictionary_results[comp_name] = comp_dict
        except:
            raise Exception("Check log error")

    def get_total_results_report(self, final_report):
        try:
            final_report_workbook = openpyxl.load_workbook(final_report)
            active_sheet = final_report_workbook[self.sheet_final_report]
            '''
            we will need to search of the name of the component based on our dictionary we have so far
            then if we have a match we will need to get the result from AO and split from "/"
            '''
            max_row = active_sheet.max_row
            needed_columns = ["C", "AO"]
            '''
            C = compo
            AO = errors/arg
            '''
            for i in range(1, max_row + 1):
                if str(active_sheet["{}{}".format(needed_columns[0], i)].value) in self.dictionary_results:
                    report_dict = {}
                    comp = str(active_sheet["{}{}".format(needed_columns[0], i)].value)
                    value_report = self.get_value_argumented(active_sheet["{}{}".format(needed_columns[1], i)].value)
                    # now just extend the dict
                    report_dict[self.list_report_items[2]] = value_report
                    self.dictionary_results[comp].update(report_dict)
        except:
            raise Exception("Check log error")

    def get_component_name(self, full_component_name):
        needed_comp_name = full_component_name.split("Impl")[0]
        return needed_comp_name

    def get_value_argumented(self, value):
        # the polyspace report must be with no errors, so we can split and take each part -> THEY MUST BE IDENTICAL
        if value is None or value == "NA":
            raise Exception("Value cannot be NA or missing - MISTAKE IN REPORT")
        needed_values = value.split("/")
        return needed_values[1]  # argumentations

    def associate_missing_parts(self):
        '''
        This functions just check if we have component and iface and just put NA AND 0 if something is missing
        LibErrHdl situation
        :return: update dictionary
        '''
        for key, value in self.dictionary_results.items():
            if not (self.list_report_items[0] in value or self.list_report_items[1] in value):
                dict_comp = {}
                dict_interface = {}
                dict_comp.update({"component": "NA"})
                dict_interface.update({"iface": "NA"})
                self.dictionary_results[key].update(dict_comp)
                self.dictionary_results[key].update(dict_interface)
            elif not (self.list_report_items[0] in value):
                dict_comp = {}
                dict_comp.update({"component": "NA"})
                self.dictionary_results[key].update(dict_comp)
            elif not (self.list_report_items[1] in value):
                dict_interface = {}
                dict_interface.update({"iface": "NA"})
                self.dictionary_results[key].update(dict_interface)

    def check_results(self):
        '''
        this will basically check if we have the same result the sum of the component and iface with the results from total report
        :return: update dictionary
        '''
        for comp, value in self.dictionary_results.items():
            result_comp = value[self.list_report_items[0]]
            result_iface = value[self.list_report_items[1]]
            result_report = value[self.list_report_items[2]]
            result_check = ""
            if value[self.list_report_items[0]] == "NA":
                result_comp = 0
            if value[self.list_report_items[1]] == "NA":
                result_iface = 0
            # make check now
            if int(result_comp) + int(result_iface) == int(result_report):
                result_check = "PASSED"
            else:
                result_check = "FAILED"
            # update dictionary
            temp_dict = {}
            temp_dict.update({self.list_report_items[3]: result_check})
            self.dictionary_results[comp].update(temp_dict)
        print(self.dictionary_results)

    def create_report(self):
        string_report = ""
        for comp, value in self.dictionary_results.items():
            string_report += "------------------------" + comp + "-----------------------\n"
            string_report += self.list_report_items[0] + " : " + str(value[self.list_report_items[0]]) + "\n"
            string_report += self.list_report_items[1] + " : " + str(value[self.list_report_items[1]]) + "\n"
            string_report += "TLT REPORT" + " : " + str(value[self.list_report_items[2]]) + "\n"
            check = value[self.list_report_items[3]]
            result_check = ""
            if check == "PASSED":
                result_check = "PASSED\t" + "✔️\n"
            else:
                result_check = "FAILED\t" + "❌\n"
            string_report += result_check + "\n\n"

        if  not os.path.exists(self.report_folder):
            os.makedirs(self.report_folder)

        with open(file=os.path.join(self.report_folder, "report_checker.txt"), mode="w", encoding="utf-8") as report_file:
            report_file.write(string_report)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Introduce the location of the 2 excel files as arguments: 1 - report polyspace\t 2 - final report")
        exit(1)
    report = ReportVerifier()
    report.get_results_polyspace_report(sys.argv[1])
    report.get_total_results_report(sys.argv[2])
    report.associate_missing_parts()
    report.check_results()
    report.create_report()
